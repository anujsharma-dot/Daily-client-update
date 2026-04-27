from flask import Flask, request, jsonify
from flask_cors import CORS
import pandas as pd
import json
import re
import io
import os
import base64
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})

PASSWORD = "Vishal@1234mumbai"

def read_file(file_storage, sheet_name=None):
    name = file_storage.filename.lower()
    if name.endswith('.csv'):
        return pd.read_csv(file_storage, dtype=str).fillna('')
    else:
        return pd.read_excel(file_storage, dtype=str, sheet_name=sheet_name or 0)

def read_all_sheets(file_storage):
    return pd.read_excel(file_storage, sheet_name=None, dtype=str)

def find_col(df, candidates):
    cols_lower = {c.lower().strip(): c for c in df.columns}
    for c in candidates:
        if c.lower().strip() in cols_lower:
            return cols_lower[c.lower().strip()]
    return None

def safe_float(val):
    try:
        return float(str(val).replace(',', '').strip())
    except:
        return 0.0

def get_working_days_in_month(year, month, day):
    count = 0
    for d in range(1, day):
        if datetime(year, month, d).weekday() < 5:
            count += 1
    return count

def assign_week(date_val, week_config):
    if not date_val or pd.isna(date_val) or str(date_val).strip() == '':
        return ''
    try:
        if isinstance(date_val, str):
            date_val = pd.to_datetime(date_val, dayfirst=False, errors='coerce')
        if pd.isna(date_val):
            return ''
        wd = get_working_days_in_month(date_val.year, date_val.month, date_val.day)
        for wk in week_config:
            if wd >= wk['start'] and wd <= wk['end']:
                return f"Week {wk['week']}"
        return ''
    except:
        return ''

def match_source_group(json_str, source_map):
    if not json_str or str(json_str).strip() in ('', 'nan'):
        return 'Others'
    s = str(json_str).lower()
    for rule in source_map:
        has_all = all(k.lower() in s for k in rule['keywords'])
        has_none = all(k.lower() not in s for k in rule['not_keywords'])
        if has_all and has_none:
            return rule['source_group']
    return 'Others'

def build_source_map(source_df):
    rules = []
    kw_col = find_col(source_df, ['Prospect Source Details Name (Keywords)', 'Keywords', 'keyword'])
    sg_col = find_col(source_df, ['Source Group', 'source_group'])
    if not kw_col or not sg_col:
        return rules
    for _, row in source_df.iterrows():
        kw_str = str(row.get(kw_col, ''))
        sg = str(row.get(sg_col, '')).strip()
        if not kw_str or not sg or sg == 'nan':
            continue
        keywords = [m.group(1) for m in re.finditer(r'Contains\s+"([^"]+)"', kw_str, re.IGNORECASE)]
        not_keywords = [m.group(1) for m in re.finditer(r'Does not contains?\s+"([^"]+)"', kw_str, re.IGNORECASE)]
        if keywords:
            rules.append({'keywords': keywords, 'not_keywords': not_keywords, 'source_group': sg})
    return rules

IP_PRODUCT_MAP = {
    'mutual fund': 'MF-SIP', 'mutual fund ': 'MF-SIP', 'sip': 'MF-SIP', 'sip ': 'MF-SIP',
    'mutual funds cob': 'MF-SIP', 'mutual funds': 'MF-SIP',
    'bond': 'Bonds', 'bonds': 'Bonds', 'bonds ': 'Bonds',
    'corporate bonds': 'Bonds', 'bonds 54ec': 'Bonds',
    'pms': 'PMS', 'pms ': 'PMS', 'pms-top up': 'PMS',
}

def get_ip_product_type(row):
    pt_col = str(row.get('Product Type', row.get('Product 2', ''))).strip()
    if pt_col in ('MF - SIP', 'Bonds', 'PMS'):
        return {'MF - SIP': 'MF-SIP', 'Bonds': 'Bonds', 'PMS': 'PMS'}[pt_col]
    return IP_PRODUCT_MAP.get(str(row.get('Product', '')).strip().lower(), None)

def process_margin(df, margin_type='gross'):
    t = margin_type
    days = ['05', '9', '15', '22', '30', '60', 'as_on']
    beyond_cash = 'Cash Margin Beyound 60' if t == 'gross' else 'Net_cash_Margin beyound 60'
    beyond_stock = 'Stock Margin Beyound 60' if t == 'gross' else 'Net_stock_Margin beyound 60'
    for d in days:
        label = 'As_On' if d == 'as_on' else f'T+{d}'
        cash = df[f'cash_{d}_days_{t}_margin'].apply(safe_float) if f'cash_{d}_days_{t}_margin' in df.columns else pd.Series(0, index=df.index)
        stock = df[f'stock_{d}_days_{t}_margin'].apply(safe_float) if f'stock_{d}_days_{t}_margin' in df.columns else pd.Series(0, index=df.index)
        w = cash + stock * 0.2
        df[f'Total {label} Weighted {t.title()} Margin'] = w
        df[f'Total {label} Capped {t.title()} Margin'] = w.apply(lambda x: min(x, 1_000_000))
    cb = df[beyond_cash].apply(safe_float) if beyond_cash in df.columns else pd.Series(0, index=df.index)
    sb = df[beyond_stock].apply(safe_float) if beyond_stock in df.columns else pd.Series(0, index=df.index)
    wb_val = cb + sb * 0.2
    df[f'Total Beyond60 Weighted {t.title()} Margin'] = wb_val
    df[f'Total Beyond60 Capped {t.title()} Margin'] = wb_val.apply(lambda x: min(x, 1_000_000))
    cc_col = find_col(df, ['Client Code', 'clientcode', 'ClientCode'])
    t60_cap_col = f'Total T+60 Capped {t.title()} Margin'
    margin_map = {}
    if cc_col and t60_cap_col in df.columns:
        for _, row in df.iterrows():
            cc = str(row[cc_col]).strip()
            if cc:
                margin_map[cc] = safe_float(row[t60_cap_col]) / 100_000
    return df, margin_map

def make_excel_b64(df, sheet_name, highlight_cols=None):
    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = list(df.columns)
    yellow = PatternFill('solid', fgColor='FFFF00')
    bold = Font(bold=True)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        if highlight_cols and any(k in h for k in highlight_cols):
            cell.fill = yellow
            cell.font = bold
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()

@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'version': '2.0.0'})

@app.route('/api/auth', methods=['POST'])
def auth():
    data = request.get_json()
    if data.get('password') == PASSWORD:
        return jsonify({'ok': True})
    return jsonify({'ok': False}), 401

@app.route('/api/process', methods=['POST'])
def process():
    try:
        week_config = json.loads(request.form.get('weekConfig', '[]'))
        action = request.form.get('action', 'master')
        files = request.files
        cm_df = read_file(files['clientMaster'])
        gm_df = read_file(files['grossMargin'])
        nm_df = read_file(files['netMargin'])
        rev_df = read_file(files['revenue'])
        vol_df = read_file(files['volume'])
        sm_sheets = read_all_sheets(files['sourceMaster'])
        sm_df = next((sdf for sname, sdf in sm_sheets.items() if any('keyword' in c.lower() or 'source group' in c.lower() for c in sdf.columns)), list(sm_sheets.values())[0])
        tpp_sheets = read_all_sheets(files['tpp'])
        vol_cc = find_col(vol_df, ['Clientcode', 'ClientCode', 'Client Code'])
        vol_map = {}
        if vol_cc:
            for _, r in vol_df.iterrows():
                cc = str(r[vol_cc]).strip()
                vol_map[cc] = vol_map.get(cc, 0) + safe_float(r.get('Volume', 0))
        rev_cc = find_col(rev_df, ['clientcode', 'ClientCode', 'Client Code'])
        rev_map = {}
        if rev_cc:
            for _, r in rev_df.iterrows():
                cc = str(r[rev_cc]).strip()
                rev_map[cc] = rev_map.get(cc, 0) + safe_float(r.get('Net Revenue', r.get('net revenue', 0)))
        gm_processed, gm_map = process_margin(gm_df, 'gross')
        nm_processed, nm_map = process_margin(nm_df, 'net')

        def agg_sheet(sheet_df, cc_candidates, amt_cols):
            cc_c = find_col(sheet_df, cc_candidates)
            amt_c = find_col(sheet_df, amt_cols)
            result = {}
            if not cc_c: return result
            for _, r in sheet_df.iterrows():
                cc = str(r[cc_c]).strip()
                if not cc or cc == 'nan': continue
                amt = safe_float(r[amt_c]) if amt_c else 0
                if cc not in result: result[cc] = {'amount': 0, 'count': 0}
                result[cc]['amount'] += amt
                result[cc]['count'] += 1
            return result

        def find_tpp(keywords):
            for sname in tpp_sheets:
                if any(k.lower() in sname.lower() for k in keywords):
                    return tpp_sheets[sname]
            return list(tpp_sheets.values())[0]

        iap_map  = agg_sheet(find_tpp(['IAP']), ['ClientCode','Client Code'], ['InvestmentAmount','Investment Amount'])
        tgs_map  = agg_sheet(find_tpp(['TGS']), ['ClientCode','Client Code'], ['TGS Amt','TGS Amount'])
        ssp_map  = agg_sheet(find_tpp(['SSP']), ['Clientcode','ClientCode','Client Code'], ['Amount'])
        tm_map   = agg_sheet(find_tpp(['Teji','TM']), ['clientCode','ClientCode','Client Code'], ['Invested_Amount','InvestmentAmount'])
        r360_map = agg_sheet(find_tpp(['R360','Research']), ['Mosl Client ID','ClientCode','Client Code'], ['amount','Amount'])
        ip_df = find_tpp(['IP'])
        ip_cc = find_col(ip_df, ['Client Code','ClientCode'])
        mfsip_map, bonds_map, pms_map = {}, {}, {}
        if ip_cc:
            for _, r in ip_df.iterrows():
                cc = str(r[ip_cc]).strip()
                if not cc or cc == 'nan': continue
                pt = get_ip_product_type(r)
                amt = safe_float(r.get('Amount', 0))
                target = {'MF-SIP': mfsip_map, 'Bonds': bonds_map, 'PMS': pms_map}.get(pt)
                if target is not None:
                    if cc not in target: target[cc] = {'amount': 0, 'count': 0}
                    target[cc]['amount'] += amt
                    target[cc]['count'] += 1
        source_map = build_source_map(sm_df)
        cm_cc = find_col(cm_df, ['Client Code','ClientCode','clientcode'])
        cm_act_date = find_col(cm_df, ['Account Activation Date','Activation Date'])
        cm_act_month = find_col(cm_df, ['Account Activation Month','Activation Month'])
        cm_lead_src = find_col(cm_df, ['Lead Source','lead_source'])
        cm_lead_src_det = find_col(cm_df, ['Lead Source Details','lead_source_details'])
        cm_unit = find_col(cm_df, ['UNIT','Unit'])
        cm_diy = find_col(cm_df, ['DIY/Non DIY','Non DIY','IS DIY','UNIT DIT/NON DIY'])
        master_rows = []
        for _, row in cm_df.iterrows():
            cc = str(row[cm_cc]).strip() if cm_cc else ''
            if not cc or cc == 'nan': continue
            act_date = row[cm_act_date] if cm_act_date else ''
            act_month = row[cm_act_month] if cm_act_month else ''
            lead_src_raw = str(row[cm_lead_src_det]) if cm_lead_src_det else ''
            lead_src = str(row[cm_lead_src]) if cm_lead_src else ''
            vol = vol_map.get(cc, 0) / 100_000
            gm = gm_map.get(cc, 0)
            nm = nm_map.get(cc, 0)
            iap = iap_map.get(cc, {'amount':0,'count':0})
            tgs = tgs_map.get(cc, {'amount':0,'count':0})
            ssp = ssp_map.get(cc, {'amount':0,'count':0})
            tm = tm_map.get(cc, {'amount':0,'count':0})
            r360 = r360_map.get(cc, {'amount':0,'count':0})
            mfs = mfsip_map.get(cc, {'amount':0,'count':0})
            bnd = bonds_map.get(cc, {'amount':0,'count':0})
            pms = pms_map.get(cc, {'amount':0,'count':0})
            ip_amt = mfs['amount'] + bnd['amount'] + pms['amount']
            ip_cnt = mfs['count'] + bnd['count'] + pms['count']
            total_products = sum(1 for c in [ssp['count'],tm['count'],tgs['count'],iap['count'],ip_cnt,r360['count']] if c > 0)
            master_rows.append({
                'Client Code': cc,
                'Account Activation Date': str(act_date),
                'Account Activation Month': str(act_month),
                'Lead Source': lead_src or match_source_group(lead_src_raw, source_map),
                'UNIT': str(row[cm_unit]) if cm_unit else '',
                'DIY/Non DIY': str(row[cm_diy]) if cm_diy else '',
                'Week': assign_week(act_date, week_config),
                'SSP Amount': round(ssp['amount'],2), 'SSP IDs': ssp['count'],
                'TM Amount': round(tm['amount'],2), 'TM IDs': tm['count'],
                'TGS Amount': round(tgs['amount'],2), 'TGS IDs': tgs['count'],
                'IMP/IAP Amount': round(iap['amount'],2), 'IMP/IAP IDs': iap['count'],
                'IP Amount': round(ip_amt,2), 'IP IDs': ip_cnt,
                'MF SIP Amount': round(mfs['amount'],2), 'MF SIP IDs': mfs['count'],
                'Bonds Amount': round(bnd['amount'],2), 'Bonds IDs': bnd['count'],
                'PMS Amount': round(pms['amount'],2), 'PMS IDs': pms['count'],
                'R360 Amount': round(r360['amount'],2), 'R360 IDs': r360['count'],
                'First Trade Done': 1 if vol > 0 else 0,
                'T+60 Gross Margin': round(gm,4),
                'T+60 Net Margin': round(nm,4),
                'Till Date Revenue': round(rev_map.get(cc,0)/100_000,4),
                'Total Products': total_products,
                'Acs with Min 1 Product': 1 if total_products >= 1 else 0,
                'MA >25K (T+60 G)': 1 if gm*100_000 > 25000 else 0,
                'MA >25K (T+60 N)': 1 if nm*100_000 > 25000 else 0,
                'Volume': round(vol,4),
            })
        master_df = pd.DataFrame(master_rows)
        if action == 'master':
            return jsonify({'file': make_excel_b64(master_df, 'Master Raw Data'), 'filename': 'Master_Raw_Data.xlsx'})
        elif action == 'gross':
            return jsonify({'file': make_excel_b64(gm_processed, 'Gross Margin Processed', ['Weighted Gross Margin','Capped Gross Margin']), 'filename': 'Gross_Margin_Processed.xlsx'})
        elif action == 'net':
            return jsonify({'file': make_excel_b64(nm_processed, 'Net Margin Processed', ['Weighted Net Margin','Capped Net Margin']), 'filename': 'Net_Margin_Processed.xlsx'})
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
